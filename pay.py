# -*- coding:UTF-8 -*-

import MergeExcel as mxl
import openpyxl as xl
import time

today = time.strftime('%y%m%d', time.localtime())


# 报销
def reimburse():
    mxl.merge(srcDir='报销', nHead=0, maxCol=3, dstDir=f'{today}-七部报销汇总.xlsx')

    ws = xl.load_workbook(f'{today}-七部报销汇总.xlsx').active
    bxSum = sum(ws['a':'a'])
    print(bxSum)


def directPay():
    srcSheet = mxl.merge('三级池', '三级池汇总.tmp.xlsx', nHead=4, nFoot=1, keepHead=False,
                         keepFoot=False, rename=True)
    # read specific columns
    '''三级池
    支付编号        2
    项目名称：      3
    供应商名称：    4
    开户银行名称：  5
    银行账号：      6
    收方银行类型    7
    收方开户行省    8
    收方开户行市    9
    本次付款金额：  11
    '''

    nRow = srcSheet.max_row
    print(f'Opened src book with {nRow} lines.')

    remark = srcSheet.iter_columns(2, 2, values_only=True)
    proj = srcSheet.iter_columns(3, 3, values_only=True)
    name = srcSheet.iter_columns(4, 4, values_only=True)
    bankName = srcSheet.iter_columns(5, 5, values_only=True)
    bankNo = srcSheet.iter_columns(6, 6, values_only=True)
    bankType = srcSheet.iter_columns(7, 7, values_only=True)
    pv = srcSheet.iter_columns(8, 8, values_only=True)
    city = srcSheet.iter_columns(9, 9, values_only=True)
    money = srcSheet.iter_columns(11, 11, values_only=True)

    print('Loaded columns.')
    # to summary workbook
    '''支付汇总表
    序号            //row
    项目名称        //3     proj
    收款单位名称    //4     name
    开户银行名称    //5     bankName
    银行账号        //6     bankNo
    本次付款金额    //11    money
    备注            //2     remark
    '''
    sumBook = xl.load_workbook('资金支付汇总表.xlsx')
    sumSheet = sumBook.active

    no = list(range(1, nRow + 1))
    z = zip(no, proj, name, bankName, bankNo, money, remark)
    content = [
        n + p + na + bna + bno + m + r
        for n, p, na, bna, bno, m, r in z
        ]
    print('Generated sum content, adding...')
    for r in content:
        sumSheet.append(r)
    sumSheet.move_range('a6:g7', nRow + 2)
    sumSheet.delete_rows(6, 2)
    sumSheet['G4'] = f'=SUM(F6:F{6 + nRow})'
    sumBook.save('资金支付汇总表.xlsx')
    print('Saved sumBook.')
    # to CBS book

    '''支付模板
    业务类型        202
    付方账号        531900085010917
    收方账号        //6                 bankNo
    收方名称        //4                 name
    收方开户行      //5                 bankName
    收方银行类型    //7                 bankType
    收方开户行省    //8                 pv
    收方开户行市    //9                 city
    交易金额        //11                money
    用途            //2                 remark
                    //nl*4
    结算方式        2
    支付渠道        3
    业务摘要        //2                 remark
    期望日          //today()20190201   today()
    期望时间        000000
    是否加急        N
    是否同城        0
                    //nl*6
    对公对私标志    2
                    //nl*2
    '''
    bankBook = xl.load_workbook('集中支付模板.xlsx')
    bankSheet = bankBook.active

    nl = ['']
    module1 = [['202', "'021900172410612"] for _ in range(nRow)]
    module2 = [bno + n + bna + bt + p + c + m + r
               for bno, n, bna, bt, p, c, m, r
               in zip(bankNo, name, bankName, bankType, pv, city, money, remark)
               ]
    module3 = [nl * 4 + [2, 3] for _ in range(nRow)]
    # __ = remark
    module4 = [[today, '000000', 'N', '0'] + nl * 6 + ['2'] + nl * 2 for _ in range(nRow)]

    content = [m1 + m2 + m3 + r + m4 for m1, m2, m3, r, m4 in
               zip(module1, module2, module3, remark, module4)]
    print('Generated bank content, adding...')

    for r in content:
        bankSheet.append(r)

    bankBook.save('集中支付模板.xlsx')
    print('Saved bankBook.')


if __name__ == '__main__':
    directPay()