# -*- coding: UTF-8 -*-

import os
import shutil
# import time

import openpyxl as pyxl
import progressbar as p
import xlrd


def main():
    srcDir = input("Insert path where the excel files to be merged exist: ")
    if not os.path.isdir(srcDir):
        raise FileNotFoundError('Not a valid path', srcDir)
    dstFile = input("Insert path where to put the dst file, or where the dst file is: ")
    pattern = input(
        'Insert nHead lines and nFoot lines lines with comma as delimiter("2,2" or "2,'
        '" or ",2"): ').replace(
        '，', ',')
    # print(f'{pattern=}')
    nHead, nFoot = [int(i) if i else 0 for i in pattern.split(',')] if pattern else (0, 0)
    merge(srcDir=srcDir, dstDir=dstFile, nHead=nHead, nFoot=nFoot)


def merge(srcDir: str = None, dstDir: str = None, nHead: int = 0, nFoot: int = 0,
          extraArea: str = None, minCol: int = None, maxCol: int = None, rename: bool = False,
          keepHead: bool = True, keepFoot: bool = True):
    """
    :param srcDir: Path where the excel files to be merged exist.
    :param dstDir: Path where to put the dst file, or where the dst file is
    :param nHead: Head line number
    :param nFoot: nFoot line number
    :param extraArea: Additional area in nHead, nFooter or content to give extra information
    with comma as delimiter.
    Use *name for filename.
    :parm minCol: start column to merge, from 1 insteading of 0
    :parm maxCol: end column to merge, from 1
    :parm rename: Whether to rename the file merged
    """

    # loop dir
    xlFiles = []
    nPass = 0
    for root, dirs, files in os.walk(srcDir):
        for f in files:
            # EXT = f.split('.')[-1].lower()
            # path = os.path.join(root, f)
            if "xls" in f.split('.')[-1].lower():
                xlFiles.append(os.path.join(root, f))
            else:
                nPass += 1

    if xlFiles:
        nXlFiles = len(xlFiles)
        maxDirLen = len(max(xlFiles, key=len))
        print(
            f'Establish dir tree, found {nXlFiles} files to be merged and {nPass} non-Excel '
            f'files passed.')
    else:
        raise FileNotFoundError(srcDir, 'has no file.')

    # open dst Excel
    # Got nothing, copy the last src file to ./ and rename to merge.xlsx as a dst
    # Got a name, copy the last src file to ./ and rename to it as a dst
    # Got a dir, copy the last src file to there and rename to merge.xlsx as a dst
    # Got a file, raise error

    if not dstDir:
        dstDir = os.path.join('./', 'merge.xlsx')
    elif os.path.isdir(dstDir):
        dstDir = os.path.join(dstDir, 'mergedExcel.xlsx')
    elif os.path.isfile(dstDir):
        os.remove(dstDir)
    else:
        dstDir = os.path.join('./', dstDir)

    specFile = xlFiles.pop()
    try:
        shutil.copy2(specFile, dstDir)
    except shutil.SameFileError:
        os.remove(dstDir)
        shutil.copy2(specFile, dstDir)

    # Load dst workbook.
    print(f'Loading dst workbook {dstDir}...')
    dstBook = pyxl.load_workbook(dstDir)
    dstSheet = dstBook.active
    nTargetRow = dstSheet.max_row
    # Copy nFoot line and paste when all src files were merged.
    if keepFoot:
        footContent = dstSheet.iter_rows(min_row=nTargetRow - nFoot + 1, values_only=True)
        dstSheet.delete_rows(nTargetRow - nFoot + 1, amount=nFoot)
    else:
        footContent = []

    if not keepHead and nHead != 0:
        dstSheet.delete_rows(1, amount=nHead)
    elif not keepHead and nHead == 0:
        raise ValueError(
            'Head num not specified, do_not_keep_head function cannot work properly.')
    dstBook.save(dstDir)

    # loop src Excel
    w = [
        p.Percentage(), ' ', p.Counter(), f'/{nXlFiles}',
        p.Bar(left=" |", right='| ', marker=">", fill="-"),
        p.ETA(), ' ', p.Variable('Progressing', width=maxDirLen, precision=maxDirLen)
        ]
    # if extraArea != '*name':
    #     # noinspection PyUnresolvedReferences
    #     extraArea = xl.utils.range_to_tuple('sheet1!' + extraArea)[:1]
    with p.ProgressBar(widgets=w, max_value=nXlFiles, redirect_stdout=True) as bar:
        for xl in xlFiles:
            bar.update(value=xlFiles.index(xl), Progressing=xl)

            if xl.split('.')[-1] == 'xls':
                newRow = readXls(xl, minCol=minCol, maxCol=maxCol, nHead=nHead, nFoot=nFoot)
            else:
                newRow = readXlsx(xl, minCol=minCol, maxCol=maxCol, nHead=nHead, nFoot=nFoot)
            for row in newRow:
                # newRow = [cell.value for cell in row]
                dstSheet.append(row)
            if rename:
                os.rename(xl, xl.replace('.xls', '-done.xls'))
            # time.sleep(1)

    for row in footContent:
        dstSheet.append(row)
    # time.sleep(0.2)
    # sys.stdout.flush()
    print(f'Finished all src files; saving to {dstDir}...')
    dstBook.save(dstDir)
    return dstSheet


def readXls(srcFp, nHead, nFoot, minCol=None, maxCol=None):
    minCol = minCol - 1 if minCol else None
    maxCol = maxCol - 1 if maxCol else None

    srcBook = xlrd.open_workbook(srcFp)
    srcSheet = srcBook.sheet_by_index(0)
    nSrcRow = srcSheet.nrows
    content = [
        srcSheet.row_values(rowx=i, start_colx=minCol, end_colx=maxCol)
        for i in range(nHead, nSrcRow - nFoot)
        ]
    return content


def readXlsx(srcFp, nHead, nFoot, minCol=None, maxCol=None):
    srcBook = pyxl.load_workbook(srcFp, data_only=True)
    srcSheet = srcBook.active
    nSrcRow = srcSheet.max_row
    content = srcSheet.iter_rows(min_row=nHead + 1, max_row=nSrcRow - nFoot, values_only=True,
                                 min_col=minCol, max_col=maxCol)
    return content


if __name__ == '__main__':
    main()
